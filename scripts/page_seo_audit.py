from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl

SITE_ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = SITE_ROOT.parent
APP_ROOT = SITE_ROOT / "app" / "[locale]"
OUTPUT_DIR = SITE_ROOT / "reports"
TOPICAL_MAP_PATH = PROJECT_ROOT / "topical-map" / "FeetToPixels-Topical-Map.xlsx"
KEYWORD_RESEARCH_PATH = (
    PROJECT_ROOT / "keyword-research" / "FeetToPixels-Keyword-Research-Master.xlsx"
)

MANUAL_PAGES: dict[str, dict[str, str]] = {
    "home": {
        "page_type": "homepage",
        "primary_keyword": "pixel conversion tools",
        "secondary_keywords": "pixel converter, dpi calculator, pixels to inches",
        "canonical_query": "pixel conversion tools",
        "focus_keyword": "pixel conversion tools",
        "title": "Pixel Conversion Tools: 15+ DPI Calculators [2026]",
        "description": (
            "Pixel conversion tools for inches, cm, mm, feet, and CSS units. "
            "Compare 15+ DPI-aware calculators, PPI guides, and print/web size "
            "references for designers."
        ),
        "source": "manual",
    },
    "about": {
        "page_type": "about",
        "primary_keyword": "about feettopixels",
        "secondary_keywords": "pixel conversion tools, dpi-aware converters",
        "canonical_query": "about feettopixels",
        "focus_keyword": "about feettopixels",
        "title": "About FeetToPixels - DPI-Aware Pixel Conversion Tools",
        "description": (
            "Learn how FeetToPixels helps designers and developers convert pixels, "
            "inches, centimeters, millimeters, feet, and CSS units with practical "
            "DPI-aware tools."
        ),
        "source": "manual",
    },
    "contact": {
        "page_type": "contact",
        "primary_keyword": "contact feettopixels",
        "secondary_keywords": "pixel converter support, dpi calculator feedback",
        "canonical_query": "contact feettopixels",
        "focus_keyword": "contact feettopixels",
        "title": "Contact FeetToPixels - Tool Support, Feedback, and Bug Reports",
        "description": (
            "Contact the FeetToPixels team for tool feedback, bug reports, content "
            "suggestions, or partnership questions related to our pixel conversion tools."
        ),
        "source": "manual",
    },
    "privacy": {
        "page_type": "legal",
        "primary_keyword": "feettopixels privacy policy",
        "secondary_keywords": "privacy policy, cookies, analytics",
        "canonical_query": "feettopixels privacy policy",
        "focus_keyword": "feettopixels privacy policy",
        "title": "FeetToPixels Privacy Policy - Data, Cookies, and Analytics",
        "description": (
            "Read the FeetToPixels privacy policy to understand how we handle contact "
            "messages, cookies, analytics, advertising, and browser-based tool activity."
        ),
        "source": "manual",
    },
    "terms": {
        "page_type": "legal",
        "primary_keyword": "feettopixels terms of service",
        "secondary_keywords": "terms of service, use of conversion tools",
        "canonical_query": "feettopixels terms of service",
        "focus_keyword": "feettopixels terms of service",
        "title": "FeetToPixels Terms of Service - Use of Our Conversion Tools",
        "description": (
            "Review the FeetToPixels terms of service for rules, limitations, and "
            "general conditions related to our pixel conversion tools and site content."
        ),
        "source": "manual",
    },
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_topical_registry() -> dict[str, dict[str, str]]:
    workbook = openpyxl.load_workbook(TOPICAL_MAP_PATH, read_only=True, data_only=True)
    seo_sheet = workbook["SEO Metadata"]
    brief_sheet = workbook["Content Briefs"]

    seo_headers = [clean(cell) for cell in next(seo_sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    seo_index = {header: index for index, header in enumerate(seo_headers)}
    brief_headers = [clean(cell) for cell in next(brief_sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    brief_index = {header: index for index, header in enumerate(brief_headers)}

    briefs_by_title: dict[str, dict[str, str]] = {}
    for row in brief_sheet.iter_rows(min_row=2, values_only=True):
        title = clean(row[brief_index["Page Title"]])
        if not title:
          continue
        briefs_by_title[title] = {
            "primary_keyword": clean(row[brief_index["PRIMARY KEYWORD"]]),
            "secondary_keywords": clean(row[brief_index["SECONDARY KEYWORDS"]]),
            "canonical_query": clean(row[brief_index["CANONICAL QUERY"]]),
            "focus_keyword": clean(row[brief_index["RANKMATH FOCUS KW"]]),
            "extractive_answer": clean(row[brief_index["EXTRACTIVE ANSWER (40 words)"]]),
            "page_type": clean(row[brief_index["CONTENT TYPE"]]).lower() or "page",
        }

    registry: dict[str, dict[str, str]] = {}
    for row in seo_sheet.iter_rows(min_row=2, values_only=True):
        raw_slug = clean(row[seo_index["URL Slug"]])
        if not raw_slug:
            continue
        slug = raw_slug.strip("/")
        title = clean(row[seo_index["Page Title"]])
        merged = {
            "title": title,
            "description": clean(row[seo_index["Meta Description"]]),
            "source": "topical-map",
        }
        merged.update(briefs_by_title.get(title, {}))
        registry[slug] = merged

    return registry


def load_keyword_research_validation() -> dict[str, dict[str, str]]:
    workbook = openpyxl.load_workbook(KEYWORD_RESEARCH_PATH, read_only=True, data_only=True)
    cluster_sheet = workbook["Clusters"]
    headers = [clean(cell) for cell in next(cluster_sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {header: position for position, header in enumerate(headers)}

    clusters: dict[str, dict[str, str]] = {}
    for row in cluster_sheet.iter_rows(min_row=2, values_only=True):
        cluster_id = clean(row[index["Cluster ID"]])
        if not cluster_id:
            continue
        clusters[cluster_id] = {
            "cluster_top_keyword": clean(row[index["Top Keyword"]]),
            "aeo_potential": clean(row[index["AEO Potential"]]),
        }

    return clusters


def get_route_slugs() -> list[str]:
    slugs = ["home"]
    for path in APP_ROOT.iterdir():
        if path.is_dir() and (path / "page.tsx").exists():
            slugs.append(path.name)
    return ["home"] + sorted(slug for slug in slugs if slug != "home")


def get_page_files(slug: str) -> list[Path]:
    if slug == "home":
        return [APP_ROOT / "page.tsx"]
    return sorted((APP_ROOT / slug).rglob("*.tsx"))


def get_page_path(slug: str) -> Path:
    if slug == "home":
        return APP_ROOT / "page.tsx"
    return APP_ROOT / slug / "page.tsx"


def read_combined_text(paths: list[Path]) -> str:
    chunks: list[str] = []
    for path in paths:
        chunks.append(path.read_text(encoding="utf-8", errors="ignore"))
    return "\n".join(chunks)


def get_page_type(slug: str, text: str, fallback: str) -> str:
    if slug in MANUAL_PAGES:
        return MANUAL_PAGES[slug]["page_type"]
    if "ConverterLayout" in text:
        return "tool"
    if "BlogLayout" in text:
        return "guide"
    if "StructuredDoc" in text:
        return "page"
    return fallback or "page"


def compute_score(slug: str, page_text: str, combined_text: str, data: dict[str, str]) -> int:
    score = 0

    if data.get("primary_keyword"):
        score += 20

    if slug == "home":
        if "englishHomeTitle" in page_text:
            score += 15
        if "englishHomeDescription" in page_text:
            score += 15
        if "pixel conversion" in combined_text.lower():
            score += 10
    else:
        if f'"{slug}"' in page_text and ".title" in page_text:
            score += 15
        if f'"{slug}"' in page_text and ".description" in page_text:
            score += 15
        if "extractive" in combined_text:
            score += 10

    if "generateMetadata" in page_text and "alternates" in page_text and "openGraph" in page_text:
        score += 10

    if "ConverterLayout" in combined_text or "BlogLayout" in combined_text:
        score += 20
    elif "DocSectionNav" in combined_text or "StructuredDoc" in combined_text:
        score += 15

    if "FAQ" in combined_text or "faqItems" in combined_text or "doc.faq" in combined_text:
        score += 5

    if "JsonLd" in combined_text:
        score += 10
    if "EditorialMeta" in combined_text or "dateModified" in combined_text or "lastUpdated" in combined_text:
        score += 5
    if "Breadcrumbs" in combined_text or "getBreadcrumbs" in combined_text:
        score += 5

    if "getRelatedTools" in combined_text or "relatedArticles" in combined_text:
        score += 10
    elif "Link" in combined_text:
        score += 5

    if "ReferenceSources" in combined_text or "DocSectionNav" in combined_text or "tocItems" in combined_text:
        score += 5

    return min(score, 100)


def build_registry() -> list[dict[str, Any]]:
    topical_registry = load_topical_registry()
    keyword_research = load_keyword_research_validation()
    page_seo_text = (SITE_ROOT / "lib" / "page-seo.ts").read_text(encoding="utf-8")

    rows: list[dict[str, Any]] = []

    for slug in get_route_slugs():
        files = get_page_files(slug)
        page_text = get_page_path(slug).read_text(encoding="utf-8", errors="ignore")
        combined_text = read_combined_text(files)

        data = dict(topical_registry.get(slug, {}))
        if slug in MANUAL_PAGES:
            data.update(MANUAL_PAGES[slug])

        if slug != "home" and slug in page_seo_text:
            data["seo_override"] = "yes"
        elif slug == "home":
            data["seo_override"] = "homepage constants"
        else:
            data["seo_override"] = "no"

        research = keyword_research.get(slug, {})
        page_type = get_page_type(slug, combined_text, data.get("page_type", ""))
        score = compute_score(slug, page_text, combined_text, data)

        rows.append(
            {
                "slug": slug,
                "page_type": page_type,
                "main_keyword": data.get("primary_keyword", ""),
                "secondary_keywords": data.get("secondary_keywords", ""),
                "canonical_query": data.get("canonical_query", ""),
                "focus_keyword": data.get("focus_keyword", ""),
                "title": data.get("title", ""),
                "description": data.get("description", ""),
                "source": data.get("source", "topical-map"),
                "research_validation": research.get("cluster_top_keyword", ""),
                "aeo_potential": research.get("aeo_potential", ""),
                "seo_override": data.get("seo_override", "unknown"),
                "score": score,
                "status": "pass" if score >= 80 else "needs work",
            }
        )

    return rows


def write_report(rows: list[dict[str, Any]]) -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)

    markdown_lines = [
        "# Page SEO Audit",
        "",
        "This audit uses the topical map workbook for page metadata and primary keywords,",
        "then validates each mapped route against the keyword research clusters where available.",
        "",
        "Scoring note: this is a local heuristic score based on keyword mapping, metadata,",
        "extractive copy, schema, sectioning, internal links, and trust signals. It is not a",
        "third-party SEO tool score.",
        "",
        "| Slug | Type | Main Keyword | Secondary Keyword | Canonical Query | Research Check | Score | Status |",
        "| --- | --- | --- | --- | --- | --- | --- | --- |",
    ]

    for row in rows:
        markdown_lines.append(
            "| {slug} | {page_type} | {main_keyword} | {secondary_keywords} | "
            "{canonical_query} | {research_validation} | {score} | {status} |".format(**row)
        )

    json_path = OUTPUT_DIR / "page-seo-audit.json"
    md_path = OUTPUT_DIR / "page-seo-audit.md"
    json_path.write_text(json.dumps(rows, indent=2), encoding="utf-8")
    md_path.write_text("\n".join(markdown_lines) + "\n", encoding="utf-8")


def main() -> None:
    rows = build_registry()
    write_report(rows)
    passing = sum(1 for row in rows if row["score"] >= 80)
    print(f"Audited {len(rows)} pages. {passing} pages are at or above the local 80-point baseline.")


if __name__ == "__main__":
    main()
